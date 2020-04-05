import openpyxl, smtplib, sys, os

em = os.environ.get('MAIL_USERNAME')
pw = os.environ.get('MAIL_PASSWORD')

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']

lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value
print(latestMonth) # 'Jun 2014'

unpaidMembers = {}
# skip first row (header)
for r in range(2, sheet.max_row + 1):
    cellV = sheet.cell(row=r, column=lastCol).value
    if cellV != 'paid':
        name = sheet.cell(r, column=1).value
        email = sheet.cell(r, column=2).value
        unpaidMembers[name] = email

print(unpaidMembers)

smtpO = smtplib.SMTP('smtp.gmail.com', 587)
smtpO.starttls()
smtpO.login(em, pw)

for name, email in unpaidMembers.items():
    body = f"Subject: {latestMonth} dues unpaid.\nDear {name}, \nRecords show that you have not paid dues for {latestMonth}. Please make payment as soon as possible. Thank you!"
    print(f"Sending email to {email}")
    try:
        smtpO.sendmail('dues@email.com', email, body)
    except:
        raise

smtpO.quit()