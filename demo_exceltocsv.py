import os
import openpyxl
import csv

files = [f for f in os.listdir('reports') if f.endswith('.xlsx')]
os.makedirs('csv', exist_ok=True)

for f in files:
    wb = openpyxl.load_workbook(os.path.join('reports', f))
    sheets = wb.sheetnames
    for sheet in sheets:
        currentsheet = wb[sheet]
        csvName = os.path.splitext(f)[0] + '-' + sheet + '.csv'
        print(csvName)
        with open(os.path.join('csv', csvName), 'w', newline='') as csvF:
            writer = csv.writer(csvF)
            for rowNum in range(1, currentsheet.max_row+1):
                rowData = []
                for colNum in range(1, currentsheet.max_column+1):
                    rowData.append(currentsheet.cell(row=rowNum, column=colNum).value)
                writer.writerow(rowData)

